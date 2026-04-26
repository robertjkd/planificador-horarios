from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.template.loader import render_to_string
from django.utils import timezone
from datetime import datetime, timedelta
from academico.models import (
    AnioAcademico,
    Asignacion,
    FranjaHoraria,
    Grupo,
    Profesor,
    Local,
    AsignacionProfesor,
    Auditoria,
    TipoAccionAuditoria,
)
from academico.utils import registrar_auditoria
from .forms import EditarAsignacionForm
from planificacion.permissions import planificador_required


def formatear_celda(asignacion):
    """
    Formatea el contenido de una celda de horario.
    
    Para Educación Física (tipo 'E'): solo abreviatura de asignatura.
    Para otras actividades: abreviatura-tipo-local.
    """
    abrev = asignacion.actividad_plan.asignatura.abreviatura
    tipo = asignacion.actividad_plan.tipo_actividad
    local = asignacion.local.codigo
    
    if tipo == 'E':
        return abrev
    return f"{abrev}-{tipo}-{local}"


@login_required
def ver_horario(request):
    """
    Vista principal para visualizar el horario académico.
    
    Parámetros GET:
    - anio_id: año académico (obligatorio)
    - semana: número de semana (obligatorio)
    - grupo_id: filtro por grupo (opcional)
    - profesor_id: filtro por profesor (opcional)
    """
    # Obtener parámetros
    anio_id = request.GET.get('anio_id')
    semana = request.GET.get('semana')
    grupo_id = request.GET.get('grupo_id')
    profesor_id = request.GET.get('profesor_id')
    
    if not anio_id or not semana:
        # Si no hay parámetros, mostrar formulario de selección
        annos = AnioAcademico.objects.all()
        return render(request, 'horario/seleccionar_horario.html', {
            'annos': annos,
            'semanas': range(1, 14),
        })
    
    anio = get_object_or_404(AnioAcademico, pk=anio_id)
    semana = int(semana)
    
    # Obtener grupos del año
    grupos = list(anio.grupos.all())
    
    # Filtrar por grupo si se especifica y el usuario no es planificador
    if grupo_id:
        grupo_seleccionado = get_object_or_404(Grupo, pk=grupo_id)
        # Solo permitir ver el propio grupo si es estudiante
        if not (request.user.es_planificador or request.user.es_vicedecano):
            # Aquí podrías agregar lógica para verificar que el grupo pertenece al estudiante
            pass
        grupos = [grupo_seleccionado]
    else:
        grupo_seleccionado = None
    
    # Obtener franjas horarias del turno del año (ordenadas por hora)
    franjas = list(FranjaHoraria.objects.filter(
        turno=anio.turno
    ).order_by('hora_inicio'))
    
    # Obtener asignaciones de la semana y año
    asignaciones_qs = Asignacion.objects.filter(
        actividad_plan__semana=semana,
        actividad_plan__asignatura__anio=anio,
    ).select_related(
        'actividad_plan__asignatura',
        'actividad_plan__grupo',
        'local',
        'franja',
        'profesor',
    )
    
    # Filtrar por profesor si se especifica
    if profesor_id:
        asignaciones_qs = asignaciones_qs.filter(profesor_id=profesor_id)
    
    # Filtrar por grupo si se especifica, pero incluir conferencias
    if grupo_id:
        asignaciones_qs = asignaciones_qs.filter(
            actividad_plan__grupo__isnull=True  # Conferencias
        ) | asignaciones_qs.filter(
            actividad_plan__grupo_id=grupo_id  # Actividades del grupo
        )
    
    asignaciones = list(asignaciones_qs)
    
    # Organizar datos en estructura jerárquica
    # { dia_semana: { franja_id: { grupo_id: asignacion } } }
    horario = {}
    for dia in range(5):  # 0-4 (Lun-Vie)
        horario[dia] = {}
        for f in franjas:
            # Las franjas aplican a todos los días
            horario[dia][f.pk] = {g.pk: None for g in grupos}
    
    # Llenar con asignaciones
    for asig in asignaciones:
        # El día viene de la actividad planificada, no de la franja
        dia = asig.actividad_plan.dia_semana if asig.actividad_plan else 0
        franja_id = asig.franja.pk if asig.franja else None
        grupo = asig.actividad_plan.grupo if asig.actividad_plan else None
        
        if franja_id is None or dia not in horario:
            continue
        if franja_id not in horario[dia]:
            continue
            
        if grupo is None:
            # Conferencia: asignar a todos los grupos
            for g in grupos:
                horario[dia][franja_id][g.pk] = asig
        else:
            # Actividad de grupo específico
            if grupo.pk in horario[dia][franja_id]:
                horario[dia][franja_id][grupo.pk] = asig
    
    # Organizar franjas por día para el template
    # Las franjas son las mismas para todos los días
    franjas_por_dia = {}
    for dia in range(5):
        franjas_por_dia[dia] = franjas
    
    # Nombres de días
    nombres_dias = {0: 'LUNES', 1: 'MARTES', 2: 'MIÉRCOLES', 3: 'JUEVES', 4: 'VIERNES'}
    
    # Obtener profesores para el filtro
    profesores = list(Profesor.objects.all())
    
    context = {
        'anio': anio,
        'semana': semana,
        'grupos': grupos,
        'grupo_seleccionado': grupo_seleccionado,
        'franjas_por_dia': franjas_por_dia,
        'horario': horario,
        'formatear_celda': formatear_celda,
        'profesor_id': profesor_id,
        'nombres_dias': nombres_dias,
        'dias': range(5),
        'annos': AnioAcademico.objects.all(),
        'semanas': range(1, 14),
        'profesores': profesores,
    }
    
    return render(request, 'horario/horario.html', context)


def construir_datos_exportacion(anio, semana, grupo_id=None, profesor_id=None):
    """
    Construye la estructura de datos tabulares para exportación.
    
    Retorna un diccionario con:
    - dias: lista de días (0-4)
    - nombres_dias: mapeo de número a nombre
    - grupos: lista de grupos
    - franjas_por_dia: {dia: [franja1, franja2, ...]}
    - horario: {dia: {franja_id: {grupo_id: asignacion}}}
    """
    # Obtener grupos del año
    grupos = list(anio.grupos.all())
    
    if grupo_id:
        grupo_seleccionado = get_object_or_404(Grupo, pk=grupo_id)
        grupos = [grupo_seleccionado]
    else:
        grupo_seleccionado = None
    
    # Obtener franjas horarias del turno del año
    franjas = list(FranjaHoraria.objects.filter(
        turno=anio.turno
    ).order_by('hora_inicio'))
    
    # Obtener asignaciones de la semana y año
    asignaciones_qs = Asignacion.objects.filter(
        actividad_plan__semana=semana,
        actividad_plan__asignatura__anio=anio,
    ).select_related(
        'actividad_plan__asignatura',
        'actividad_plan__grupo',
        'local',
        'franja',
        'profesor',
    )
    
    # Filtrar por profesor si se especifica
    if profesor_id:
        asignaciones_qs = asignaciones_qs.filter(profesor_id=profesor_id)
    
    # Filtrar por grupo si se especifica, pero incluir conferencias
    if grupo_id:
        asignaciones_qs = asignaciones_qs.filter(
            actividad_plan__grupo__isnull=True
        ) | asignaciones_qs.filter(
            actividad_plan__grupo_id=grupo_id
        )
    
    asignaciones = list(asignaciones_qs)
    
    # Organizar datos en estructura jerárquica
    horario = {}
    for dia in range(5):
        horario[dia] = {}
        for f in franjas:
            # Las franjas aplican a todos los días
            horario[dia][f.pk] = {g.pk: None for g in grupos}
    
    # Llenar con asignaciones
    for asig in asignaciones:
        # El día viene de la actividad planificada
        dia = asig.actividad_plan.dia_semana if asig.actividad_plan else 0
        franja_id = asig.franja.pk if asig.franja else None
        grupo = asig.actividad_plan.grupo if asig.actividad_plan else None
        
        if franja_id is None or dia not in horario:
            continue
        if franja_id not in horario[dia]:
            continue
            
        if grupo is None:
            # Conferencia: asignar a todos los grupos
            for g in grupos:
                horario[dia][franja_id][g.pk] = asig
        else:
            # Actividad de grupo específico
            if grupo.pk in horario[dia][franja_id]:
                horario[dia][franja_id][grupo.pk] = asig
    
    # Organizar franjas por día (las mismas para todos los días)
    franjas_por_dia = {}
    for dia in range(5):
        franjas_por_dia[dia] = franjas
    
    # Nombres de días
    nombres_dias = {0: 'LUNES', 1: 'MARTES', 2: 'MIÉRCOLES', 3: 'JUEVES', 4: 'VIERNES'}
    
    return {
        'anio': anio,
        'semana': semana,
        'grupos': grupos,
        'grupo_seleccionado': grupo_seleccionado,
        'franjas_por_dia': franjas_por_dia,
        'horario': horario,
        'nombres_dias': nombres_dias,
        'dias': range(5),
    }


def generar_excel(datos):
    """Genera un archivo Excel con el horario usando openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Horario"
    
    # Estilos
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    center_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row = 1
    
    # Para cada día
    for dia in datos['dias']:
        if not datos['franjas_por_dia'].get(dia):
            continue
        
        # Título del día
        ws.merge_cells(f'{get_column_letter(1)}{row}:{get_column_letter(len(datos["grupos"]) + 1)}{row}')
        cell = ws.cell(row=row, column=1, value=datos['nombres_dias'][dia])
        cell.font = Font(bold=True, size=14)
        cell.alignment = center_alignment
        row += 1
        
        # Encabezados de columna
        ws.cell(row=row, column=1, value="Hora").font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=1).alignment = center_alignment
        
        for col, grupo in enumerate(datos['grupos'], start=2):
            cell = ws.cell(row=row, column=col, value=grupo.nombre)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        row += 1
        
        # Filas de franjas
        for franja in datos['franjas_por_dia'][dia]:
            # Hora
            hora_texto = f"{franja.hora_inicio.strftime('%H:%M')}–{franja.hora_fin.strftime('%H:%M')}"
            cell = ws.cell(row=row, column=1, value=hora_texto)
            cell.alignment = center_alignment
            cell.border = thin_border
            
            # Celdas de grupos
            for col, grupo in enumerate(datos['grupos'], start=2):
                asignacion = datos['horario'][dia].get(franja.pk, {}).get(grupo.pk)
                if asignacion:
                    cell_value = formatear_celda(asignacion)
                else:
                    cell_value = ""
                
                cell = ws.cell(row=row, column=col, value=cell_value)
                cell.alignment = center_alignment
                cell.border = thin_border
            
            row += 1
        
        row += 1  # Espacio entre días
    
    # Ajustar anchos de columna
    for col in range(1, len(datos['grupos']) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Guardar en memoria
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def generar_csv(datos):
    """Genera un archivo CSV con el horario."""
    import csv
    from io import StringIO
    
    output = StringIO()
    writer = csv.writer(output)
    
    # Cabeceras
    writer.writerow(['Día', 'Hora', 'Grupo', 'Asignatura', 'Tipo', 'Local', 'Profesor'])
    
    # Para cada día y franja
    for dia in datos['dias']:
        if not datos['franjas_por_dia'].get(dia):
            continue
        
        nombre_dia = datos['nombres_dias'][dia]
        
        for franja in datos['franjas_por_dia'][dia]:
            hora_texto = f"{franja.hora_inicio.strftime('%H:%M')}–{franja.hora_fin.strftime('%H:%M')}"
            
            for grupo in datos['grupos']:
                asignacion = datos['horario'][dia].get(franja.pk, {}).get(grupo.pk)
                
                if asignacion:
                    actividad = asignacion.actividad_plan
                    asignatura = actividad.asignatura
                    tipo = actividad.tipo_actividad
                    local = asignacion.local.codigo
                    profesor = asignacion.profesor.nombre if asignacion.profesor else "Sin asignar"
                    
                    # Para Educación Física, tipo es vacío en la celda
                    if tipo == 'E':
                        celda_texto = asignatura.abreviatura
                    else:
                        celda_texto = f"{asignatura.abreviatura}-{tipo}-{local}"
                    
                    writer.writerow([
                        nombre_dia,
                        hora_texto,
                        grupo.nombre,
                        asignatura.nombre,
                        tipo,
                        local,
                        profesor
                    ])
    
    output.seek(0)
    return output


def generar_pdf(datos):
    """Genera un PDF con el horario usando xhtml2pdf."""
    from django.template.loader import render_to_string
    from xhtml2pdf import pisa
    from io import BytesIO
    
    # Renderizar template HTML
    html = render_to_string('horario/pdf_horario.html', {
        'datos': datos,
        'formatear_celda': formatear_celda,
    })
    
    # Convertir a PDF
    output = BytesIO()
    pisa.CreatePDF(html, dest=output)
    output.seek(0)
    
    return output


@login_required
def exportar_horario(request):
    """
    Vista para exportar el horario en diferentes formatos (PDF, Excel, CSV).
    
    Parámetros GET:
    - anio_id: año académico (obligatorio)
    - semana: número de semana (obligatorio)
    - grupo_id: filtro por grupo (opcional)
    - profesor_id: filtro por profesor (opcional)
    - formato: 'pdf', 'excel', o 'csv' (obligatorio)
    """
    anio_id = request.GET.get('anio_id')
    semana = request.GET.get('semana')
    grupo_id = request.GET.get('grupo_id')
    profesor_id = request.GET.get('profesor_id')
    formato = request.GET.get('formato', 'pdf')
    
    if not anio_id or not semana:
        from django.contrib import messages
        messages.error(request, 'Debe seleccionar año y semana para exportar.')
        return redirect('horario:ver_horario')
    
    anio = get_object_or_404(AnioAcademico, pk=anio_id)
    semana = int(semana)
    
    # Restricción para usuarios de consulta
    if request.user.es_consulta and not (request.user.es_planificador or request.user.es_vicedecano):
        # Si es estudiante, forzar su grupo (si tiene grupo asignado)
        if request.user.grupo:
            grupo_id = request.user.grupo.pk
        # Si es profesor, forzar sus asignaciones (lógica adicional podría ir aquí)
    
    # Construir datos de exportación
    datos = construir_datos_exportacion(anio, semana, grupo_id, profesor_id)
    
    # Generar archivo según formato
    if formato == 'excel':
        output = generar_excel(datos)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="horario_semana_{semana}.xlsx"'
        return response
    
    elif formato == 'csv':
        output = generar_csv(datos)
        response = HttpResponse(
            output.getvalue(),
            content_type='text/csv'
        )
        response['Content-Disposition'] = f'attachment; filename="horario_semana_{semana}.csv"'
        return response
    
    elif formato == 'pdf':
        output = generar_pdf(datos)
        response = HttpResponse(
            output.read(),
            content_type='application/pdf'
        )
        response['Content-Disposition'] = f'attachment; filename="horario_semana_{semana}.pdf"'
        return response
    
    else:
        from django.contrib import messages
        messages.error(request, 'Formato no válido.')
        return redirect('horario:ver_horario')


# ─────────────────────────────────────────────────────────────
# API ENDPOINTS FOR FULLCALENDAR
# ─────────────────────────────────────────────────────────────

@login_required
def api_horario(request):
    """
    API endpoint para obtener eventos del horario en formato JSON para FullCalendar.
    
    Parámetros GET:
    - anio_id: año académico (obligatorio)
    - semana: número de semana (obligatorio)
    - grupo_id: filtro por grupo (opcional)
    - profesor_id: filtro por profesor (opcional)
    """
    anio_id = request.GET.get('anio_id')
    semana = request.GET.get('semana')
    grupo_id = request.GET.get('grupo_id')
    profesor_id = request.GET.get('profesor_id')
    
    if not anio_id or not semana:
        return JsonResponse({'error': 'anio_id y semana son obligatorios'}, status=400)
    
    anio = get_object_or_404(AnioAcademico, pk=anio_id)
    semana = int(semana)
    
    # Calcular fecha base para la semana (lunes de esa semana)
    # Asumimos que semana 1 es la primera semana del año académico
    # Para simplificar, usamos una fecha base fija y calculamos offset
    fecha_base = datetime(2026, 1, 1)  # Ajustar según año académico real
    offset_semanas = semana - 1
    fecha_lunes = fecha_base + timedelta(weeks=offset_semanas)
    fecha_lunes = fecha_lunes - timedelta(days=fecha_lunes.weekday())  # Ajustar a lunes
    
    # Obtener asignaciones
    asignaciones_qs = Asignacion.objects.filter(
        actividad_plan__semana=semana,
        actividad_plan__asignatura__anio=anio,
    ).select_related(
        'actividad_plan__asignatura',
        'actividad_plan__grupo',
        'local',
        'franja',
        'profesor',
    )
    
    # Filtrar por profesor
    if profesor_id:
        asignaciones_qs = asignaciones_qs.filter(profesor_id=profesor_id)
    
    # Filtrar por grupo (pero incluir conferencias)
    if grupo_id:
        asignaciones_qs = asignaciones_qs.filter(
            actividad_plan__grupo__isnull=True
        ) | asignaciones_qs.filter(
            actividad_plan__grupo_id=grupo_id
        )
    
    asignaciones = asignaciones_qs.all()
    
    # Colores por tipo de actividad
    colores_tipo = {
        'C': '#e3f2fd',      # Conferencia - Azul claro
        'CE': '#bbdefb',     # Conferencia especial - Azul
        'CP': '#c8e6c9',     # Clase práctica - Verde claro
        'L': '#c5e1a5',      # Laboratorio - Verde
        'S': '#fff9c4',      # Seminario - Amarillo claro
        'T': '#ffe0b2',      # Taller - Naranja claro
        'TE': '#ffcc80',     # Taller especial - Naranja
        'PP': '#f5f5f5',     # Práctica profesional - Gris claro
        'E': '#f8bbd0',      # Educación física - Rosa
        'NP': '#e0e0e0',     # No presencial - Gris
    }
    
    eventos = []
    for asig in asignaciones:
        actividad = asig.actividad_plan
        asignatura = actividad.asignatura
        grupo = actividad.grupo
        franja = asig.franja
        
        # Calcular fecha y hora del evento
        # El día viene de la actividad planificada, no de la franja
        dia_offset = asig.actividad_plan.dia_semana if asig.actividad_plan else 0
        fecha_evento = fecha_lunes + timedelta(days=dia_offset)
        
        # Combinar fecha con hora de inicio y fin
        hora_inicio = datetime.combine(fecha_evento.date(), franja.hora_inicio)
        hora_fin = datetime.combine(fecha_evento.date(), franja.hora_fin)
        
        # Formatear título
        tipo = actividad.tipo_actividad
        local = asig.local.codigo
        
        if tipo == 'E':
            titulo = asignatura.abreviatura
        else:
            titulo = f"{asignatura.abreviatura}-{tipo}-{local}"
        
        # Añadir indicador de grupo si es conferencia
        if grupo is None:
            titulo += " (Todos)"
        else:
            titulo += f" ({grupo.nombre})"
        
        # Construir evento
        evento = {
            'id': asig.pk,
            'asignacion_id': asig.pk,
            'title': titulo,
            'start': hora_inicio.isoformat(),
            'end': hora_fin.isoformat(),
            'backgroundColor': colores_tipo.get(tipo, '#ffffff'),
            'extendedProps': {
                'tipo_actividad': tipo,
                'grupo_nombre': grupo.nombre if grupo else None,
                'grupo_id': grupo.pk if grupo else None,
                'asignatura_nombre': asignatura.nombre,
                'asignatura_abreviatura': asignatura.abreviatura,
                'profesor_nombre': asig.profesor.nombre if asig.profesor else 'Sin asignar',
                'profesor_id': asig.profesor.pk if asig.profesor else None,
                'local_codigo': local,
                'local_id': asig.local.pk,
                'franja_id': franja.pk,
                'semana': semana,
            }
        }
        
        eventos.append(evento)
    
    return JsonResponse(eventos, safe=False)


@login_required
@require_http_methods(["POST"])
def api_modificar_asignacion(request, asignacion_id):
    """
    API endpoint para modificar una asignación vía AJAX.
    
    Recibe JSON con: franja_id, local_id, profesor_id
    Devuelve JSON con success true/false y errores si los hay.
    """
    try:
        asignacion = get_object_or_404(Asignacion, pk=asignacion_id)
    except:
        return JsonResponse({'success': False, 'errors': {'__all__': ['Asignación no encontrada']}}, status=404)
    
    # Verificar permisos
    if not (request.user.es_planificador or request.user.es_vicedecano):
        return JsonResponse({'success': False, 'errors': {'__all__': ['No tiene permiso para modificar asignaciones']}}, status=403)
    
    # Obtener datos del request
    import json
    try:
        data = json.loads(request.body)
    except:
        return JsonResponse({'success': False, 'errors': {'__all__': ['JSON inválido']}}, status=400)
    
    franja_id = data.get('franja_id')
    local_id = data.get('local_id')
    profesor_id = data.get('profesor_id')
    
    # Validar que al menos se cambie algo
    if not franja_id and not local_id and not profesor_id:
        return JsonResponse({'success': False, 'errors': {'__all__': ['Debe especificar al menos un campo a modificar']}}, status=400)
    
    # Actualizar campos si se proporcionan
    if franja_id:
        try:
            nueva_franja = FranjaHoraria.objects.get(pk=franja_id)
            asignacion.franja = nueva_franja
        except FranjaHoraria.DoesNotExist:
            return JsonResponse({'success': False, 'errors': {'franja': ['Franja no válida']}}, status=400)
    
    if local_id:
        try:
            nuevo_local = Local.objects.get(pk=local_id)
            asignacion.local = nuevo_local
        except Local.DoesNotExist:
            return JsonResponse({'success': False, 'errors': {'local': ['Local no válido']}}, status=400)
    
    if profesor_id:
        if profesor_id == 'null' or profesor_id == '':
            asignacion.profesor = None
        else:
            try:
                nuevo_profesor = Profesor.objects.get(pk=profesor_id)
                asignacion.profesor = nuevo_profesor
            except Profesor.DoesNotExist:
                return JsonResponse({'success': False, 'errors': {'profesor': ['Profesor no válido']}}, status=400)
    
    # Validar usando el formulario
    form = EditarAsignacionForm(data, instance=asignacion)
    if not form.is_valid():
        errors = {}
        for field, field_errors in form.errors.items():
            errors[field] = list(field_errors)
        return JsonResponse({'success': False, 'errors': errors}, status=400)
    
    # Guardar
    asignacion.save()
    
    return JsonResponse({'success': True, 'message': 'Asignación actualizada correctamente'})


@login_required
def api_franjas(request):
    """
    API endpoint para obtener franjas horarias filtradas por año.
    """
    anio_id = request.GET.get('anio_id')
    
    if not anio_id:
        return JsonResponse({'error': 'anio_id es obligatorio'}, status=400)
    
    anio = get_object_or_404(AnioAcademico, pk=anio_id)
    
    franjas = FranjaHoraria.objects.filter(
        turno=anio.turno
    ).order_by('hora_inicio')
    
    data = []
    for f in franjas:
        data.append({
            'id': f.pk,
            'orden': f.orden,
            'hora_inicio': f.hora_inicio.strftime('%H:%M'),
            'hora_fin': f.hora_fin.strftime('%H:%M'),
            'turno': f.turno,
        })
    
    return JsonResponse(data, safe=False)


@login_required
def api_locales(request):
    """
    API endpoint para obtener locales compatibles.
    
    Parámetros:
    - tipo_actividad: tipo de actividad (opcional)
    - capacidad_necesaria: capacidad mínima (opcional)
    """
    tipo_actividad = request.GET.get('tipo_actividad')
    capacidad_necesaria = request.GET.get('capacidad_necesaria')
    
    locales = Local.objects.all()
    
    if tipo_actividad:
        # Mapeo de tipo de actividad a tipo de local
        tipo_local_map = {
            'C': 'S',
            'CE': 'S',
            'L': 'L',
            'CP': 'A',
            'S': 'A',
            'T': 'A',
            'TE': 'A',
            'PP': 'A',
            'E': 'O',
            'NP': None,
        }
        tipo_local = tipo_local_map.get(tipo_actividad)
        if tipo_local:
            locales = locales.filter(tipo=tipo_local)
        elif tipo_actividad == 'NP':
            locales = Local.objects.none()
    
    if capacidad_necesaria:
        try:
            capacidad = int(capacidad_necesaria)
            locales = locales.filter(capacidad__gte=capacidad)
        except ValueError:
            pass
    
    data = []
    for l in locales:
        data.append({
            'id': l.pk,
            'codigo': l.codigo,
            'nombre': l.nombre,
            'tipo': l.tipo,
            'capacidad': l.capacidad,
        })
    
    return JsonResponse(data, safe=False)


@login_required
def api_profesores(request):
    """
    API endpoint para obtener profesores asignados a una asignatura/grupo.
    
    Parámetros:
    - asignatura_id: ID de asignatura (opcional)
    - grupo_id: ID de grupo (opcional)
    """
    asignatura_id = request.GET.get('asignatura_id')
    grupo_id = request.GET.get('grupo_id')
    
    if not asignatura_id:
        return JsonResponse({'error': 'asignatura_id es obligatorio'}, status=400)
    
    from academico.models import Asignatura
    asignatura = get_object_or_404(Asignatura, pk=asignatura_id)
    
    if grupo_id:
        # Profesores asignados a ese grupo
        asignaciones = AsignacionProfesor.objects.filter(
            asignatura=asignatura,
            grupo_id=grupo_id
        )
    else:
        # Profesores asignados a conferencia (sin grupo)
        asignaciones = AsignacionProfesor.objects.filter(
            asignatura=asignatura,
            grupo__isnull=True
        )
    
    profesor_ids = asignaciones.values_list('profesor_id', flat=True)
    profesores = Profesor.objects.filter(id__in=profesor_ids)
    
    data = []
    for p in profesores:
        data.append({
            'id': p.pk,
            'nombre': p.nombre,
            'apellido': p.apellido if hasattr(p, 'apellido') else '',
        })
    
    return JsonResponse(data, safe=False)


@login_required
def ver_horario_calendario(request):
    """
    Vista del horario con FullCalendar interactivo.
    """
    annos = AnioAcademico.objects.all()
    semanas = range(1, 14)
    grupos = Grupo.objects.all()
    profesores = Profesor.objects.all()
    
    # Obtener filtros de la URL
    anio_id = request.GET.get('anio_id')
    semana = request.GET.get('semana', '1')
    grupo_id = request.GET.get('grupo_id')
    profesor_id = request.GET.get('profesor_id')
    
    # Determinar si el usuario puede editar
    puede_editar = request.user.es_planificador or request.user.es_vicedecano
    
    context = {
        'annos': annos,
        'semanas': semanas,
        'grupos': grupos,
        'profesores': profesores,
        'anio_id': anio_id,
        'semana': semana,
        'grupo_id': grupo_id,
        'profesor_id': profesor_id,
        'puede_editar': puede_editar,
    }
    
    return render(request, 'horario/horario_calendario.html', context)


@planificador_required
def editar_asignacion(request, asignacion_id):
    """
    Vista para editar manualmente una asignación de horario.
    
    GET: Devuelve el formulario HTML para el modal.
    POST: Procesa el formulario y devuelve JSON con resultado.
    """
    asignacion = get_object_or_404(Asignacion, pk=asignacion_id)
    
    if request.method == 'GET':
        form = EditarAsignacionForm(instance=asignacion)
        return render(request, 'horario/modal_editar_asignacion.html', {
            'form': form,
            'asignacion': asignacion,
        })
    
    elif request.method == 'POST':
        form = EditarAsignacionForm(request.POST, instance=asignacion)
        
        if form.is_valid():
            # Guardar cambios antiguos para auditoría
            cambios = []
            if 'franja' in form.changed_data:
                cambios.append(f"Franja: {asignacion.franja} → {form.cleaned_data['franja']}")
            if 'local' in form.changed_data:
                cambios.append(f"Local: {asignacion.local.codigo} → {form.cleaned_data['local'].codigo}")
            if 'profesor' in form.changed_data:
                viejo_prof = asignacion.profesor.nombre if asignacion.profesor else 'Sin asignar'
                nuevo_prof = form.cleaned_data['profesor'].nombre if form.cleaned_data['profesor'] else 'Sin asignar'
                cambios.append(f"Profesor: {viejo_prof} → {nuevo_prof}")
            
            form.save()
            
            # Registrar auditoría
            detalles = (
                f"Asignación #{asignacion.pk} modificada manualmente. "
                f"Cambios: {', '.join(cambios) if cambios else 'Ninguno'}. "
                f"Actividad: {asignacion.actividad_plan.asignatura.abreviatura}-{asignacion.actividad_plan.tipo_actividad}"
            )
            registrar_auditoria(
                request.user,
                TipoAccionAuditoria.MODIFICAR_ASIGNACION,
                detalles,
                request
            )
            
            return JsonResponse({
                'success': True,
                'message': 'Asignación actualizada correctamente.'
            })
        else:
            # Devolver errores en formato JSON
            errors = {}
            for field, field_errors in form.errors.items():
                errors[field] = list(field_errors)
            
            return JsonResponse({
                'success': False,
                'errors': errors
            }, status=400)
