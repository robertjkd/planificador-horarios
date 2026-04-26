"""
Vistas del módulo de planificación.

Incluye:
- importar_balance(): función principal de importación de balance de carga.
- importar_balance_view(): vista Django (GET/POST) con formulario.
- generar_horario_view(): ejecuta el motor OR-Tools.
- seleccionar_generacion(): formulario para elegir año y semana.
"""
import csv
import io
import logging
import os
import re
import tempfile
import uuid
from typing import BinaryIO

import pandas as pd
from openpyxl import load_workbook
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.shortcuts import get_object_or_404, redirect, render

from academico.models import (
    ActividadPlan,
    AnioAcademico,
    Asignacion,
    Asignatura,
    FranjaHoraria,
    Grupo,
    TipoActividad,
)
from academico.utils import registrar_auditoria, verificar_fallos_solver_y_alertar
from academico.models import TipoAccionAuditoria
from .forms import ImportarBalanceForm
from .permissions import vicedecano_required, planificador_required

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────
# IMPORTAR BALANCE
# ─────────────────────────────────────────────────────────────

# Mapping del CSV/Excel al modelo: 1=Lunes → 0, 2=Martes → 1, etc.
DIA_CSV_A_MODELO = {1: 0, 2: 1, 3: 2, 4: 3}

# Tipos de actividad que se crean una sola vez (sin grupo)
TIPOS_CONFERENCIA = {'C', 'CE', 'EC'}  # EC = Especial de Conferencia (código alternativo)

# Tipos de actividad que se replican por cada grupo del año
TIPOS_POR_GRUPO = {'CP', 'L', 'S', 'T', 'TE', 'E', 'PP'}

# Mapeo de códigos alternativos/normalizados a códigos oficiales
MAPEO_CODIGOS_ALTERNATIVOS = {
    'EC': 'CE',      # Especial de Conferencia → Conferencia Especial
    'CONF': 'C',     # Conferencia abreviada
    'CONFESP': 'CE', # Conferencia Especial abreviada
    'CLASE': 'CP',   # Clase → Clase Práctica
    'LAB': 'L',      # Laboratorio abreviado
    'SEM': 'S',      # Seminario abreviado
    'TALL': 'T',     # Taller abreviado
    'EDFIS': 'E',    # Educación Física abreviada
    'PRAC': 'PP',    # Práctica Profesional abreviada
    'NOPRES': 'NP',  # No Presencial abreviado
}


def _normalizar_codigo_actividad(codigo: str) -> str:
    """
    Normaliza un código de actividad a su forma oficial.
    
    Maneja variantes comunes como 'EC' → 'CE', etc.
    Si no se reconoce, devuelve el código original en mayúsculas.
    """
    if not codigo:
        return ''
    codigo = codigo.strip().upper()
    return MAPEO_CODIGOS_ALTERNATIVOS.get(codigo, codigo)


def _parsear_encabezado_matrix(encabezado: str):
    """
    Extrae semana y día de encabezados tipo 'S1D1', 'S12D4', 'Sem1Dia1', etc.
    Devuelve (semana, dia_csv) o (None, None) si no coincide.
    """
    enc = encabezado.strip().upper()
    # Patrones: S1D1, S12D4, SEM1DIA1, 1D1, etc.
    patrones = [
        r'^S(\d+)D(\d+)$',          # S1D1
        r'^SEM(\d+)DIA(\d+)$',       # SEM1DIA1
        r'^SEMANA(\d+)DIA(\d+)$',    # SEMANA1DIA1
    ]
    for patron in patrones:
        m = re.match(patron, enc)
        if m:
            return int(m.group(1)), int(m.group(2))
    return None, None


def _convertir_matrix_a_filas(df: pd.DataFrame):
    """
    Convierte un DataFrame en formato matriz (columnas S1D1, S1D2…)
    a lista de dicts con las columnas estándar:
    {'asignatura': str, 'semana': int, 'dia': int, 'actividad': str}
    """
    filas_resultado = []
    # Identificar columnas de encabezado y las que son semana/día
    col_asignatura = None
    cols_semana_dia = []  # [(nombre_columna, semana, dia), …]

    for col in df.columns:
        col_str = str(col).strip()
        if col_str.upper() in ('ASIGNATURA', 'MATERIA', 'NOMBRE'):
            col_asignatura = col
            continue
        sem, dia = _parsear_encabezado_matrix(col_str)
        if sem is not None:
            cols_semana_dia.append((col, sem, dia))

    if col_asignatura is None:
        raise ValueError(
            'No se encontró columna de asignaturas. '
            'La primera columna debe llamarse "Asignatura", "Materia" o "Nombre".'
        )

    if not cols_semana_dia:
        raise ValueError(
            'No se encontraron columnas de semana/día. '
            'Usa formato S1D1, S1D2, … o SEM1DIA1, …'
        )

    for _, row in df.iterrows():
        nombre_asig = str(row[col_asignatura]).strip()
        if not nombre_asig or nombre_asig.lower() in ('nan', 'none', ''):
            continue
        for col, semana, dia_csv in cols_semana_dia:
            valor = row[col]
            if pd.isna(valor):
                continue
            act = str(valor).strip().upper()
            if not act:
                continue
            filas_resultado.append({
                'asignatura': nombre_asig,
                'semana': semana,
                'dia': dia_csv,
                'actividad': act,
            })

    return filas_resultado


def _leer_csv(archivo: BinaryIO) -> list[dict]:
    """
    Lee un archivo CSV con columnas: asignatura, semana, dia, actividad.
    Devuelve lista de dicts normalizados.
    """
    contenido = archivo.read().decode('utf-8-sig')
    lector = csv.DictReader(io.StringIO(contenido))
    filas = []
    for row in lector:
        filas.append({
            'asignatura': str(row.get('asignatura', '')).strip(),
            'semana': int(row.get('semana', 0)),
            'dia': int(row.get('dia', 0)),
            'actividad': str(row.get('actividad', '')).strip().upper(),
        })
    return filas


def _detectar_formato_excel(archivo: BinaryIO) -> str:
    """
    Detecta el formato del archivo Excel leyendo las primeras filas.
    
    Retorna:
    - 'csv_like': Si tiene columnas ASIGNATURA, SEMANA, DIA, ACTIVIDAD
    - 'matrix': Si tiene encabezados tipo S1D1, S1D2...
    - 'real': Si tiene estructura con filas de encabezado (filas 1-3) y datos desde fila 4
    """
    try:
        df = pd.read_excel(archivo, engine='openpyxl', nrows=5)
        archivo.seek(0)  # Resetear puntero
        
        columnas = [str(c).strip().upper() for c in df.columns]
        
        # Verificar formato CSV-like
        if 'ASIGNATURA' in columnas and 'SEMANA' in columnas:
            return 'csv_like'
        
        # Verificar formato matriz (columnas tipo S1D1)
        for col in columnas:
            if re.match(r'^S\d+D\d+$', col):
                return 'matrix'
        
        # Si tiene muchas columnas sin nombres reconocibles, probablemente es formato real
        if len(columnas) > 10:
            # Verificar si la primera columna está vacía o tiene valores como "Asignatura"
            primera_col = str(df.columns[0]).strip().upper()
            if primera_col in ('', 'NAN', 'ASIGNATURA', 'MATERIA', 'NOMBRE'):
                return 'real'
        
        return 'matrix'  # Default
    except Exception:
        archivo.seek(0)
        return 'real'  # Si hay error, intentar formato real


def _leer_excel(archivo: BinaryIO) -> list[dict]:
    """
    Lee un archivo Excel (.xlsx). Detecta automáticamente el formato:
    - Formato real: filas 1-3 encabezados, fila 4+ datos (columnas B-AS, AT, AU, AV)
    - Formato matriz: encabezados S1D1, S1D2...
    - CSV-like: columnas asignatura, semana, dia, actividad
    """
    formato = _detectar_formato_excel(archivo)
    archivo.seek(0)
    
    if formato == 'real':
        return _leer_excel_formato_real(archivo)
    
    # Para otros formatos, usar pandas
    df = pd.read_excel(archivo, engine='openpyxl')
    df.columns = [str(c).strip() for c in df.columns]
    columnas_upper = [c.upper() for c in df.columns]

    # Si tiene las columnas estándar, tratar como CSV-like
    if 'ASIGNATURA' in columnas_upper and 'SEMANA' in columnas_upper:
        col_map = {}
        for c in df.columns:
            col_map[c.upper()] = c
        filas = []
        for _, row in df.iterrows():
            filas.append({
                'asignatura': str(row[col_map['ASIGNATURA']]).strip(),
                'semana': int(row[col_map['SEMANA']]),
                'dia': int(row[col_map['DIA']]),
                'actividad': str(row[col_map['ACTIVIDAD']]).strip().upper(),
            })
        return filas

    # Si no, intentar formato matriz
    return _convertir_matrix_a_filas(df)


def _leer_excel_formato_real(archivo: BinaryIO) -> list[dict]:
    """
    Lee un archivo Excel (.xlsx) con el formato específico real del balance de carga.
    
    Formato esperado:
    - Filas 1-3: Encabezados con celdas combinadas
      * Fila 1: Rangos de fechas (ej. "22 - 28 ene")
      * Fila 2: Etiquetas de semana (en col 2, 6, 10...)
      * Fila 3: Números de sesión (1,2,3,4)
    - Columna A: Nombre de la asignatura
    - Columnas B-AS (2-45): 44 columnas (11 semanas × 4 sesiones)
    - Columnas AT, AU, AV (46-48): Eventos especiales (Extra, Victoria, Mundial)
    
    Retorna lista de dicts: {'asignatura': str, 'semana': int, 'dia': int, 'actividad': str}
    """
    filas_resultado = []
    
    # Cargar workbook con openpyxl
    wb = load_workbook(archivo, data_only=True)
    hoja = wb.active
    
    # Mapeo de columnas especiales (AT=46, AU=47, AV=48 en Excel = 1-based)
    COLUMNAS_ESPECIALES = {
        46: 12,  # AT -> Semana 12 (Extra)
        47: 13,  # AU -> Semana 13 (Victoria)
        48: 14,  # AV -> Semana 14 (Mundial)
    }
    
    # Encontrar la última fila con datos
    max_fila = hoja.max_row
    max_col = min(hoja.max_column, 48)  # Limitar hasta columna AV (48)
    
    # Procesar filas de datos (desde fila 4)
    for fila_idx in range(4, max_fila + 1):
        # Obtener nombre de asignatura (columna A = 1)
        celda_asig = hoja.cell(row=fila_idx, column=1)
        nombre_asig = str(celda_asig.value).strip() if celda_asig.value else ''
        
        # Limpiar prefijos como "1." o "1. "
        if nombre_asig:
            nombre_asig = re.sub(r'^\d+\.\s*', '', nombre_asig)
        
        if not nombre_asig or nombre_asig.lower() in ('nan', 'none', ''):
            continue
        
        # Procesar columnas B-AS (2-45) -> 11 semanas × 4 sesiones
        for bloque in range(11):  # 0-10 (semanas 1-11)
            semana = bloque + 1
            for sesion in range(4):  # 0-3 (lunes-jueves)
                col_idx = 2 + (bloque * 4) + sesion  # Columna Excel (1-based)
                dia_csv = sesion + 1  # 1-4 (lunes-jueves en CSV -> 1,2,3,4)
                
                if col_idx > max_col:
                    continue
                
                celda = hoja.cell(row=fila_idx, column=col_idx)
                valor = celda.value
                
                if valor is None or str(valor).strip() == '':
                    continue
                
                codigo = str(valor).strip().upper()
                if codigo and codigo not in ('NAN', 'NONE'):
                    filas_resultado.append({
                        'asignatura': nombre_asig,
                        'semana': semana,
                        'dia': dia_csv,
                        'actividad': codigo,
                    })
        
        # Procesar columnas especiales AT, AU, AV (46-48)
        for col_idx, semana_especial in COLUMNAS_ESPECIALES.items():
            if col_idx > max_col:
                continue
            
            celda = hoja.cell(row=fila_idx, column=col_idx)
            valor = celda.value
            
            if valor is None or str(valor).strip() == '':
                continue
            
            codigo = str(valor).strip().upper()
            if codigo and codigo not in ('NAN', 'NONE'):
                # Día 0 = Lunes para eventos especiales
                filas_resultado.append({
                    'asignatura': nombre_asig,
                    'semana': semana_especial,
                    'dia': 1,  # Lunes para eventos especiales
                    'actividad': codigo,
                })
    
    wb.close()
    return filas_resultado


def importar_balance_preview(
    archivo: BinaryIO,
    anio: AnioAcademico,
    crear_asignaturas: bool = False,
    formato: str = 'auto',
) -> dict:
    """
    Genera una vista previa de la importación sin guardar en base de datos.
    
    Similar a importar_balance pero solo calcula estadísticas sin crear registros.
    
    Parámetros
    ----------
    archivo : BinaryIO
        Archivo subido (objeto File-like).
    anio : AnioAcademico
        Año académico al que pertenece el balance.
    crear_asignaturas : bool
        Si True, simula la creación de asignaturas inexistentes.
    formato : str
        'csv', 'xlsx' o 'auto' (detecta por extensión).
    
    Retorna
    -------
    dict
        {
            'success': bool,
            'anio': str,
            'asignaturas_procesadas': int,
            'actividades_creadas': int,
            'actividades_omitidas': int,
            'detalle_por_tipo': {tipo: int, ...},
            'errores': list[str],
            'filas_vacias': int,
            'advertencias': list[str],
        }
    """
    logger.info('Generando vista previa de importación para: %s', anio)
    
    grupos = list(Grupo.objects.filter(anio=anio))
    
    if not grupos:
        raise ValueError(f'El año "{anio}" no tiene grupos registrados.')
    
    # ── 1. Leer archivo ─────────────────────────────────────────
    if formato == 'auto':
        nombre = getattr(archivo, 'name', '')
        if nombre.lower().endswith('.xlsx'):
            formato = 'xlsx'
        else:
            formato = 'csv'
    
    if formato == 'xlsx':
        filas = _leer_excel(archivo)
    else:
        filas = _leer_csv(archivo)
    
    # ── 2. Procesar simulando (sin guardar) ─────────────────────
    resumen = {
        'success': True,
        'anio': str(anio),
        'asignaturas_procesadas': set(),
        'actividades_creadas': 0,
        'actividades_omitidas': 0,
        'detalle_por_tipo': {},
        'errores': [],
        'filas_vacias': 0,
        'advertencias': [],
    }
    
    # Tracking de asignaturas que se crearían y actividades que existirían
    asignaturas_simuladas = {}  # nombre -> abreviatura
    actividades_simuladas = set()  # (asignatura, tipo, grupo_id, semana, dia)
    
    for idx, fila in enumerate(filas, start=1):
        nombre_asig = fila['asignatura']
        semana = fila['semana']
        dia_csv = fila['dia']
        tipo_raw = fila['actividad']
        
        # Normalizar código de actividad
        tipo = _normalizar_codigo_actividad(tipo_raw)
        
        # Validar vacíos
        if not nombre_asig or not tipo:
            resumen['filas_vacias'] += 1
            continue
        
        # Validar día
        if dia_csv not in DIA_CSV_A_MODELO:
            resumen['errores'].append(
                f'Fila {idx}, Asignatura "{nombre_asig}": día "{dia_csv}" no válido (debe ser 1-4).'
            )
            continue
        dia_modelo = DIA_CSV_A_MODELO[dia_csv]
        
        # Validar semana
        if not (1 <= semana <= 52):
            resumen['errores'].append(
                f'Fila {idx}, Asignatura "{nombre_asig}": semana "{semana}" fuera de rango (1-52).'
            )
            continue
        
        # Buscar o simular asignatura
        try:
            asignatura = Asignatura.objects.get(nombre__iexact=nombre_asig, anio=anio)
            asig_id = asignatura.pk
            asig_nombre = asignatura.nombre
        except Asignatura.DoesNotExist:
            if crear_asignaturas:
                # Simular creación de asignatura
                abrev = nombre_asig[:10].upper().replace(' ', '')
                base_abrev = abrev
                contador = 1
                while abrev in asignaturas_simuladas.values() or Asignatura.objects.filter(abreviatura=abrev, anio=anio).exists():
                    abrev = f'{base_abrev[:8]}{contador}'
                    contador += 1
                asignaturas_simuladas[nombre_asig] = abrev
                asig_id = f'SIM_{nombre_asig}'  # ID simulado
                asig_nombre = nombre_asig
                resumen['advertencias'].append(f'Asignatura "{nombre_asig}" será creada automáticamente.')
            else:
                resumen['errores'].append(
                    f'Fila {idx}: asignatura "{nombre_asig}" no encontrada en {anio}.'
                )
                continue
        
        resumen['asignaturas_procesadas'].add(asig_nombre)
        
        # Validar tipo de actividad
        if tipo not in TipoActividad.values:
            tipos_validos = ', '.join(sorted(TipoActividad.values))
            resumen['errores'].append(
                f'Fila {idx}, Asignatura "{asig_nombre}": '
                f'tipo "{tipo_raw}" (normalizado: "{tipo}") no reconocido. '
                f'Tipos válidos: {tipos_validos}'
            )
            continue
        
        # ── Simular creación de actividades ─────────────────────
        registros_simulados = []
        
        if tipo in TIPOS_CONFERENCIA:
            registros_simulados.append({
                'asig_id': asig_id,
                'asig_nombre': asig_nombre,
                'tipo': tipo,
                'grupo_id': None,
                'semana': semana,
                'dia': dia_modelo,
            })
        elif tipo == 'NP':
            registros_simulados.append({
                'asig_id': asig_id,
                'asig_nombre': asig_nombre,
                'tipo': tipo,
                'grupo_id': None,
                'semana': semana,
                'dia': dia_modelo,
            })
        elif tipo in TIPOS_POR_GRUPO:
            for grupo in grupos:
                registros_simulados.append({
                    'asig_id': asig_id,
                    'asig_nombre': asig_nombre,
                    'tipo': tipo,
                    'grupo_id': grupo.pk,
                    'semana': semana,
                    'dia': dia_modelo,
                })
        
        # Verificar duplicados y contar
        for datos in registros_simulados:
            clave = (datos['asig_id'], datos['tipo'], datos['grupo_id'], datos['semana'], datos['dia'])
            
            # Verificar si ya existe en BD
            existe_en_bd = False
            if isinstance(datos['asig_id'], int):
                try:
                    existe_en_bd = ActividadPlan.objects.filter(
                        asignatura_id=datos['asig_id'],
                        tipo_actividad=datos['tipo'],
                        grupo_id=datos['grupo_id'],
                        anio=anio,
                        semana=datos['semana'],
                        dia_semana=datos['dia'],
                    ).exists()
                except:
                    pass
            
            if clave in actividades_simuladas or existe_en_bd:
                resumen['actividades_omitidas'] += 1
            else:
                actividades_simuladas.add(clave)
                resumen['actividades_creadas'] += 1
                resumen['detalle_por_tipo'][tipo] = resumen['detalle_por_tipo'].get(tipo, 0) + 1
    
    # Convertir set a conteo para el resumen final
    resumen['asignaturas_procesadas'] = len(resumen['asignaturas_procesadas'])
    
    if resumen['advertencias']:
        resumen['advertencias'] = list(set(resumen['advertencias']))  # Eliminar duplicados
    
    logger.info(
        'Vista previa generada: creadas=%s omitidas=%s errores=%s',
        resumen['actividades_creadas'],
        resumen['actividades_omitidas'],
        len(resumen['errores']),
    )
    return resumen


def importar_balance(
    archivo: BinaryIO,
    anio_id: int,
    user,
    formato: str = 'auto',
    limpiar_previos: bool = False,
    crear_asignaturas: bool = False,
) -> dict:
    """
    Importa un balance de carga desde un archivo CSV o Excel y crea registros
    de ``ActividadPlan``.

    Parámetros
    ----------
    archivo : BinaryIO
        Archivo subido (objeto File-like).
    anio_id : int
        PK del ``AnioAcademico`` al que pertenece el balance.
    user : User
        Usuario que ejecuta la importación (para logging/auditoría).
    formato : str
        'csv', 'xlsx' o 'auto' (detecta por extensión).
    limpiar_previos : bool
        Si True, elimina todas las ActividadPlan del año antes de importar.
    crear_asignaturas : bool
        Si True, crea asignaturas inexistentes automáticamente.

    Retorna
    -------
    dict
        {
            'success': bool,
            'anio': str,
            'asignaturas_procesadas': int,
            'actividades_creadas': int,
            'actividades_omitidas': int,
            'detalle_por_tipo': {tipo: int, ...},
            'errores': list[str],
            'filas_vacias': int,
        }
    """
    logger.info(
        'Inicio importación balance: anio_id=%s usuario=%s limpiar=%s crear_asign=%s',
        anio_id, user.username if user else 'anon', limpiar_previos, crear_asignaturas
    )

    anio = get_object_or_404(AnioAcademico, pk=anio_id)
    grupos = list(Grupo.objects.filter(anio=anio))

    if not grupos:
        raise ValueError(f'El año "{anio}" no tiene grupos registrados.')

    # ── 1. Leer archivo ─────────────────────────────────────────
    if formato == 'auto':
        # Intentar detectar por contenido; si no, asumir CSV
        nombre = getattr(archivo, 'name', '')
        if nombre.lower().endswith('.xlsx'):
            formato = 'xlsx'
        else:
            formato = 'csv'

    if formato == 'xlsx':
        filas = _leer_excel(archivo)
    else:
        filas = _leer_csv(archivo)

    # ── 2. Procesar en transacción atómica ─────────────────────
    resumen = {
        'success': True,
        'anio': str(anio),
        'asignaturas_procesadas': set(),
        'actividades_creadas': 0,
        'actividades_omitidas': 0,
        'detalle_por_tipo': {},
        'errores': [],
        'filas_vacias': 0,
    }

    with transaction.atomic():
        if limpiar_previos:
            eliminadas, _ = ActividadPlan.objects.filter(anio=anio).delete()
            logger.info('Eliminadas %s actividades previas del año %s', eliminadas, anio)

        for idx, fila in enumerate(filas, start=1):
            nombre_asig = fila['asignatura']
            semana = fila['semana']
            dia_csv = fila['dia']
            tipo_raw = fila['actividad']
            
            # Normalizar código de actividad
            tipo = _normalizar_codigo_actividad(tipo_raw)
            
            # Log para debugging de códigos no reconocidos
            if tipo_raw and tipo not in TipoActividad.values:
                logger.warning(
                    'Fila %s: código "%s" normalizado a "%s" (no reconocido como oficial)',
                    idx, tipo_raw, tipo
                )

            # Validar vacíos
            if not nombre_asig or not tipo:
                resumen['filas_vacias'] += 1
                continue

            # Validar día
            if dia_csv not in DIA_CSV_A_MODELO:
                error_msg = f'Fila {idx}, Asignatura "{nombre_asig}": día "{dia_csv}" no válido (debe ser 1-4).'
                resumen['errores'].append(error_msg)
                logger.warning(error_msg)
                continue
            dia_modelo = DIA_CSV_A_MODELO[dia_csv]

            # Validar semana
            if not (1 <= semana <= 52):
                error_msg = f'Fila {idx}, Asignatura "{nombre_asig}": semana "{semana}" fuera de rango (1-52).'
                resumen['errores'].append(error_msg)
                logger.warning(error_msg)
                continue

            # Buscar o crear asignatura
            try:
                asignatura = Asignatura.objects.get(nombre__iexact=nombre_asig, anio=anio)
            except Asignatura.DoesNotExist:
                if crear_asignaturas:
                    # Crear con abreviatura = primeras 10 chars del nombre
                    abrev = nombre_asig[:10].upper().replace(' ', '')
                    # Evitar duplicados de abreviatura en el mismo año
                    base_abrev = abrev
                    contador = 1
                    while Asignatura.objects.filter(abreviatura=abrev, anio=anio).exists():
                        abrev = f'{base_abrev[:8]}{contador}'
                        contador += 1
                    asignatura = Asignatura.objects.create(
                        nombre=nombre_asig,
                        abreviatura=abrev,
                        anio=anio,
                    )
                    logger.info('Asignatura creada automáticamente: %s', asignatura)
                else:
                    error_msg = (
                        f'Fila {idx}: asignatura "{nombre_asig}" no existe en {anio}. '
                        f'Use "Crear asignaturas inexistentes" o verifique el nombre.'
                    )
                    resumen['errores'].append(error_msg)
                    logger.warning(error_msg)
                    continue

            resumen['asignaturas_procesadas'].add(asignatura.nombre)

            # Validar tipo de actividad
            if tipo not in TipoActividad.values:
                tipos_validos = ', '.join(sorted(TipoActividad.values))
                error_msg = (
                    f'Fila {idx}, Asignatura "{nombre_asig}": '
                    f'tipo de actividad "{tipo_raw}" (normalizado: "{tipo}") no reconocido. '
                    f'Tipos válidos: {tipos_validos}'
                )
                resumen['errores'].append(error_msg)
                logger.warning(error_msg)
                continue

            # ── Crear actividades según reglas de réplica ──────────
            registros_a_crear = []

            if tipo in TIPOS_CONFERENCIA:
                # Una sola vez, sin grupo
                registros_a_crear.append({
                    'asignatura': asignatura,
                    'tipo_actividad': tipo,
                    'grupo': None,
                    'anio': anio,
                    'semana': semana,
                    'dia_semana': dia_modelo,
                    'requiere_local': True,
                })

            elif tipo == 'NP':
                # No presencial: una sola vez, sin grupo, sin local
                registros_a_crear.append({
                    'asignatura': asignatura,
                    'tipo_actividad': tipo,
                    'grupo': None,
                    'anio': anio,
                    'semana': semana,
                    'dia_semana': dia_modelo,
                    'requiere_local': False,
                })

            elif tipo in TIPOS_POR_GRUPO:
                # Replicar por cada grupo del año
                for grupo in grupos:
                    registros_a_crear.append({
                        'asignatura': asignatura,
                        'tipo_actividad': tipo,
                        'grupo': grupo,
                        'anio': anio,
                        'semana': semana,
                        'dia_semana': dia_modelo,
                        'requiere_local': True,
                    })
            else:
                # No debería llegar aquí por la validación previa
                resumen['errores'].append(
                    f'Fila {idx}: tipo "{tipo}" no tiene regla de réplica definida.'
                )
                continue

            # ── Guardar evitando duplicados ────────────────────────
            for datos in registros_a_crear:
                # Verificar si ya existe la misma combinación
                existe = ActividadPlan.objects.filter(
                    asignatura=datos['asignatura'],
                    tipo_actividad=datos['tipo_actividad'],
                    grupo=datos['grupo'],
                    anio=datos['anio'],
                    semana=datos['semana'],
                    dia_semana=datos['dia_semana'],
                ).exists()

                if existe:
                    resumen['actividades_omitidas'] += 1
                    continue

                ActividadPlan.objects.create(**datos)
                resumen['actividades_creadas'] += 1
                resumen['detalle_por_tipo'][tipo] = resumen['detalle_por_tipo'].get(tipo, 0) + 1

    # Convertir set a conteo para el resumen final
    resumen['asignaturas_procesadas'] = len(resumen['asignaturas_procesadas'])
    
    # Log detallado del desglose
    logger.info('=== RESUMEN IMPORTACIÓN ===')
    logger.info('Año: %s', resumen['anio'])
    logger.info('Asignaturas procesadas: %s', resumen['asignaturas_procesadas'])
    logger.info('Actividades creadas: %s', resumen['actividades_creadas'])
    logger.info('Actividades omitidas (duplicadas): %s', resumen['actividades_omitidas'])
    logger.info('Filas vacías: %s', resumen['filas_vacias'])
    logger.info('Errores: %s', len(resumen['errores']))
    
    if resumen['detalle_por_tipo']:
        logger.info('Desglose por tipo:')
        for tipo, count in sorted(resumen['detalle_por_tipo'].items()):
            nombre_completo = dict(TipoActividad.choices).get(tipo, tipo)
            logger.info('  %s (%s): %s actividades', tipo, nombre_completo, count)
    
    if resumen['errores']:
        logger.warning('Errores encontrados:')
        for error in resumen['errores'][:10]:  # Mostrar primeros 10
            logger.warning('  - %s', error)
        if len(resumen['errores']) > 10:
            logger.warning('  ... y %s errores más', len(resumen['errores']) - 10)
    
    logger.info('===========================')
    return resumen


@vicedecano_required
def importar_balance_view(request):
    """
    Vista para subir y procesar el archivo de balance de carga.
    
    Flujo de 2 pasos:
    1. GET → muestra formulario
    2. POST (subir archivo) → muestra vista previa sin guardar
    3. POST (confirmar) → ejecuta importación real y guarda en BD
    4. POST (cancelar) → limpia sesión y vuelve al formulario
    """
    # PASO 2: Confirmar o Cancelar la importación
    if request.method == 'POST' and request.POST.get('accion'):
        accion = request.POST.get('accion')
        
        if accion == 'cancelar':
            # Limpiar datos de sesión y volver al formulario
            request.session.pop('importacion_preview', None)
            messages.info(request, 'Importación cancelada.')
            return redirect('planificacion:importar_balance')
        
        elif accion == 'confirmar':
            # Recuperar datos de sesión y ejecutar importación real
            preview_data = request.session.get('importacion_preview')
            if not preview_data:
                messages.error(request, 'No hay datos de importación pendientes.')
                return redirect('planificacion:importar_balance')
            
            try:
                # Re-procesar el archivo temporal para importación real
                preview_data = request.session.get('importacion_preview')
                temp_path = preview_data.get('temp_file_path') if preview_data else None
                
                if not temp_path or not os.path.exists(temp_path):
                    messages.error(request, 'El archivo temporal no está disponible. Por favor, suba el archivo nuevamente.')
                    return redirect('planificacion:importar_balance')
                
                # Recuperar datos de sesión
                anio_id = preview_data['anio_id']
                limpiar_previos = preview_data.get('limpiar_previos', False)
                crear_asignaturas = preview_data.get('crear_asignaturas', False)
                
                logger.info(f"Confirmando importación: anio_id={anio_id}, limpiar={limpiar_previos}")
                
                # Ejecutar importación real
                with open(temp_path, 'rb') as archivo:
                    resumen = importar_balance(
                        archivo=archivo,
                        anio_id=anio_id,
                        user=request.user,
                        limpiar_previos=limpiar_previos,
                        crear_asignaturas=crear_asignaturas,
                    )
                
                # Limpiar sesión y archivo temporal
                request.session.pop('importacion_preview', None)
                try:
                    os.remove(temp_path)
                except OSError:
                    pass  # Ignorar error al eliminar archivo temporal
                
                # Mostrar resultado
                if resumen['errores']:
                    messages.warning(
                        request, 
                        f"Importación completada con {len(resumen['errores'])} errores. "
                        f"Actividades creadas: {resumen['actividades_creadas']}."
                    )
                else:
                    messages.success(
                        request,
                        f"¡Importación exitosa! "
                        f"Actividades creadas: {resumen['actividades_creadas']}, "
                        f"Asignaturas procesadas: {resumen['asignaturas_procesadas']}."
                    )
                
                # Registrar auditoría
                anio = AnioAcademico.objects.get(pk=anio_id)
                detalles = (
                    f"Balance importado para {anio}. "
                    f"Actividades: {resumen['actividades_creadas']}, "
                    f"Asignaturas: {resumen['asignaturas_procesadas']}."
                )
                registrar_auditoria(
                    request.user,
                    TipoAccionAuditoria.IMPORTAR_BALANCE,
                    detalles,
                    request
                )
                
                return redirect('planificacion:importar_balance')
                
            except Exception as exc:
                logger.exception('Error confirmando importación')
                messages.error(request, f'Error al confirmar la importación: {exc}')
                return redirect('planificacion:importar_balance')
    
    # PASO 1: Subir archivo y mostrar vista previa
    if request.method == 'POST':
        form = ImportarBalanceForm(request.POST, request.FILES)
        if form.is_valid():
            anio = form.cleaned_data['anio']
            archivo = request.FILES['archivo']
            limpiar = form.cleaned_data.get('limpiar_previos', False)
            crear_asignaturas = form.cleaned_data.get('crear_asignaturas', False)
            
            try:
                # Generar vista previa (sin guardar en BD)
                # Leemos el archivo y procesamos para mostrar resumen
                resumen_preview = importar_balance_preview(
                    archivo, anio, crear_asignaturas
                )
                
                # Guardar archivo temporalmente para el paso de confirmación
                temp_dir = tempfile.gettempdir()
                temp_filename = f"balance_{uuid.uuid4().hex}_{archivo.name}"
                temp_path = os.path.join(temp_dir, temp_filename)
                
                # Guardar el archivo temporalmente
                with open(temp_path, 'wb+') as temp_file:
                    for chunk in archivo.chunks():
                        temp_file.write(chunk)
                
                logger.info(f"Archivo temporal guardado: {temp_path}")
                
                # Guardar datos en sesión para el paso de confirmación
                request.session['importacion_preview'] = {
                    'anio_id': anio.pk,
                    'anio_nombre': str(anio),
                    'nombre_archivo': archivo.name,
                    'limpiar_previos': limpiar,
                    'crear_asignaturas': crear_asignaturas,
                    'resumen': resumen_preview,
                    'temp_file_path': temp_path,
                }
                
                # Preparar vista previa para mostrar
                vista_previa = {
                    'anio': str(anio),
                    'nombre_archivo': archivo.name,
                    'asignaturas_procesadas': resumen_preview['asignaturas_procesadas'],
                    'actividades_creadas': resumen_preview['actividades_creadas'],
                    'actividades_omitidas': resumen_preview['actividades_omitidas'],
                    'filas_vacias': resumen_preview['filas_vacias'],
                    'errores': resumen_preview['errores'],
                    'detalle_por_tipo': resumen_preview.get('detalle_por_tipo', {}),
                    'advertencias': resumen_preview.get('advertencias', []),
                }
                
                return render(request, 'planificacion/importar_balance.html', {
                    'form': form,
                    'mostrar_vista_previa': True,
                    'vista_previa': vista_previa,
                })
                
            except Exception as exc:
                logger.exception('Error generando vista previa')
                messages.error(request, f'Error al procesar el archivo: {exc}')
                return render(request, 'planificacion/importar_balance.html', {
                    'form': form,
                })
    else:
        form = ImportarBalanceForm()
        # Limpiar cualquier dato de sesión previo
        request.session.pop('importacion_preview', None)

    return render(request, 'planificacion/importar_balance.html', {
        'form': form,
        'mostrar_vista_previa': False,
        'mostrar_resumen': False,
    })


# ─────────────────────────────────────────────────────────────
# GENERACIÓN DE HORARIO (placeholders / scheduler)
# ─────────────────────────────────────────────────────────────

@planificador_required
def seleccionar_generacion(request):
    """
    Vista para seleccionar año y semana antes de generar el horario.
    Solo accesible para Planificador y Vicedecano.
    """
    from academico.models import AnioAcademico
    annos = AnioAcademico.objects.all()
    semanas = range(1, 17)

    if request.method == 'POST':
        anio_id = request.POST.get('anio')
        semana = request.POST.get('semana')
        if anio_id and semana:
            return redirect('planificacion:generar', anio_id=anio_id, semana=semana)

    return render(request, 'planificacion/seleccionar_generacion.html', {
        'annos': annos,
        'semanas': semanas,
    })


@planificador_required
def generar_horario_view(request, anio_id, semana):
    """
    Ejecuta el motor de planificación OR-Tools.
    Solo accesible para Planificador y Vicedecano.
    """
    from .scheduler import generar_horario
    from .horario_table import construir_tabla_horario, detectar_conflictos_profesor

    anio = get_object_or_404(AnioAcademico, pk=anio_id)
    
    tabla_horario = None
    
    try:
        resultado = generar_horario(anio_id, semana)
        
        if resultado.exito:
            # Determinar tipo de éxito (normal, relajado o emergencia)
            es_relajada = resultado.estadisticas.get('es_solucion_relajada', False)
            es_emergencia = resultado.estadisticas.get('es_solucion_emergencia', False)
            tiene_virtuales = resultado.estadisticas.get('asignaciones_virtuales', 0) > 0
            
            if es_emergencia:
                messages.error(request, resultado.mensaje)  # Error visual pero funcional
            elif es_relajada:
                messages.warning(request, resultado.mensaje)
            elif tiene_virtuales:
                messages.success(request, resultado.mensaje)
            else:
                messages.success(request, resultado.mensaje)
            
            # Construir tabla de horario para visualización
            # Incluso en emergencia, mostrar las actividades que necesitan asignación manual
            asignaciones = Asignacion.objects.filter(
                actividad_plan__semana=semana,
                actividad_plan__asignatura__anio=anio
            ).select_related(
                'actividad_plan',
                'actividad_plan__asignatura',
                'actividad_plan__grupo',
                'franja',
                'local',
                'profesor'
            )
            
            franjas = FranjaHoraria.objects.filter(turno=anio.turno)
            
            tabla_horario = construir_tabla_horario(asignaciones, franjas)
            tabla_horario = detectar_conflictos_profesor(tabla_horario)
            tabla_horario.estadisticas = resultado.estadisticas
            
            # Si es emergencia, agregar actividades pendientes a la tabla
            if es_emergencia:
                actividades_pendientes = ActividadPlan.objects.filter(
                    semana=semana,
                    asignatura__anio=anio,
                    requiere_local=True
                ).select_related('asignatura', 'grupo')
                
                # Marcar en estadísticas para el template
                resultado.estadisticas['actividades_pendientes'] = list(
                    actividades_pendientes.values(
                        'id', 'asignatura__nombre', 'grupo__nombre', 
                        'dia_semana', 'tipo_actividad'
                    )
                )
            
            # Registrar auditoría exitosa
            stats = resultado.estadisticas
            detalles = (
                f"Horario generado para semana {semana} de {anio}. "
                f"Actividades: {stats.get('actividades_asignadas', 0)}, "
                f"Tiempo: {stats.get('tiempo_segundos', 0):.2f}s, "
                f"Virtuales: {stats.get('asignaciones_virtuales', 0)}, "
                f"Relajado: {es_relajada}, "
                f"Emergencia: {es_emergencia}"
            )
            registrar_auditoria(
                request.user,
                TipoAccionAuditoria.GENERAR_HORARIO,
                detalles,
                request
            )
        else:
            messages.error(request, resultado.mensaje)
            
            # Registrar error del solver
            detalles = (
                f"Error al generar horario para semana {semana} de {anio}. "
                f"Motivo: {resultado.mensaje}"
            )
            registrar_auditoria(
                request.user,
                TipoAccionAuditoria.ERROR_SOLVER,
                detalles,
                request
            )
            
            # Verificar si hay demasiados fallos recientes y alertar
            verificar_fallos_solver_y_alertar()
            
    except Exception as e:
        logger.exception('Error en generar_horario_view')
        messages.error(request, f'Error al generar horario: {str(e)}')
        
        # Registrar error del solver
        detalles = (
            f"Excepción al generar horario para semana {semana} de {anio}. "
            f"Error: {str(e)}"
        )
        registrar_auditoria(
            request.user,
            TipoAccionAuditoria.ERROR_SOLVER,
            detalles,
            request
        )
        
        verificar_fallos_solver_y_alertar()

    return render(request, 'planificacion/generar_resultado.html', {
        'anio': anio,
        'semana': semana,
        'resultado': resultado if 'resultado' in locals() else None,
        'tabla_horario': tabla_horario,
        'dias_semana': [(0, 'Lunes'), (1, 'Martes'), (2, 'Miércoles'), (3, 'Jueves'), (4, 'Viernes')],
    })
